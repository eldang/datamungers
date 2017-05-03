#! /usr/bin/env python3

# Converts multiple-tab Excel files into flat CSVs, preserving unicode
__author__ = "Eldan Goldenberg, February 2017"
# http://eldan.co.uk/ ~ @eldang ~ eldang@gmail.com
#
# Usage: put a job list CSV like the enclosed files_to_process.csv
# into the same directory as the files it names.	Fill the columns as follows:
# filename: input file name relative to the location of the CSV itself
# header: row number (A = 1) of the row to use as a header - simply lets us skip rows if needed
# subheader: optional row number (A = 1) of a subheader row.
#		If a file has a subheader, the output CSV will have column headings in the
#		format: headervalue: subheadervalue
# tabs: column name for the text in tab names, so that their data is preserved
#		in the flat output CSV.	 Leave blank to only process the first tab
# column_wrap: optional number of columns after which the data wraps around
# special_handling: NOT IMPLEMENTED
# notes: optional human-readable notes, ignored by the script
#
# Then simply: flatten.py path/to/joblist.csv

import argparse
import csv
import openpyxl	# for newer-style .xlsx files
import os
import sys
import time
import xlrd	 		# for old-style .xls files


verbose = True
output_subdir = "flattened"




def main():
	args = get_args()
	print_with_timestamp("Starting run.")
	starttime = time.time()

	inputdir = os.path.abspath(os.path.dirname(args.joblist))
	outputdir = os.path.join(inputdir, output_subdir)
	if not os.path.isdir(outputdir):
		os.mkdir(outputdir)

	filecount = process_job_list(inputdir, outputdir, args.joblist)

	print_with_timestamp(
			"Run complete. " + str(filecount) + " files processed in "
			+ elapsed_time(starttime) + "."
	)




def process_job_list(inputdir, outputdir, joblist):
	filecount = 0
	print_if_verbose("Opening " + joblist)
	with open(joblist) as jobsfile:
		jobs = csv.DictReader(jobsfile)
		for job in jobs:
			ext = os.path.splitext(job["filename"])[1]
			job["inputfile"] = os.path.join(inputdir, job["filename"])
			outputfile = os.path.join(
					outputdir,
					os.path.basename(job["filename"]).replace(ext, ".csv")
			)
			if ext == ".xls":
				write_csv(read_xls(job), outputfile)
				filecount += 1
			elif ext == ".xlsx":
				write_csv(read_xlsx(job), outputfile)
				filecount += 1
			else:
				print("File extension " + ext + " not recognised, skipping row:")
				print(job)
	return filecount



def read_xls(job):
	print_with_timestamp("Processing " + job["filename"])
	ntabs = 0
	data = {
		'headers': [],
		'rows': []
	}
	with xlrd.open_workbook(job["inputfile"]) as workbook:
		if job["tabs"] == "":
			data = read_xls_sheet(workbook.sheet_by_index(0), data, job)
			ntabs = 1
		else:
			data["headers"].append(job["tabs"])
			for sheet in workbook.sheets():
				if job["skip_tabs"] != "":
					if job["skip_tabs"] == 1:
						job["skip_tabs"] = ""
					else:
						job["skip_tabs"] -= 1
				else:
					data = read_xls_sheet(sheet, data, job)
					ntabs += 1
	print_with_timestamp(str(ntabs) + " tab[s] read")
	return data



def read_xls_sheet(sheet, data, job):
	header = int(job["header"]) - 1 # -1 because xlrd is 0-indexed while Excel itself is 1-indexed in the UI
	if job["subheader"] == "":
		subheader = None
		firstrow = int(job["header"])
	else:
		subheader = int(job["subheader"]) - 1
		firstrow = int(job["subheader"])

	if job["column_wrap"] == "":
		ncols = sheet.ncols
		nframes = 1
	else:
		ncols = int(job["column_wrap"])
		if sheet.ncols % ncols > 0:
			nframes = sheet.ncols // ncols + 1
		else:
			nframes = sheet.ncols // ncols
	prev_head = ""
	col_names = []

	for col in range(0, ncols):
		if subheader is None:
			val = sheet.cell_value(header, col)
		else:
			if sheet.cell_value(header, col) != "":
				prev_head = sheet.cell_value(header, col)
			if sheet.cell_value(subheader, col) == "":
				val = prev_head
			else:
				val = prev_head + ": " + str(sheet.cell_value(subheader, col))
		col_names.append(val)

	for val in col_names:
		if val != "" and val not in data["headers"]:
			data["headers"].append(val)

	for frame in range(0, nframes):
		for row in range(firstrow, sheet.nrows):
			content = {}
			if job["tabs"] != "":
				content[job["tabs"]] = sheet.name
			for col in range(0, len(col_names)):
				if col_names[col] != "":
					content[col_names[col]] = clean_value(sheet.cell_value(row, col + frame * ncols), True)
			data["rows"].append(content)

	return data



def read_xlsx(job):
	print_with_timestamp("Processing " + job["filename"])
	ntabs = 0
	data = {
		'headers': [],
		'rows': []
	}
	wb = openpyxl.load_workbook(
		job["inputfile"],
		read_only=True,
		keep_vba=False,
		guess_types=True,
		data_only=True,
		keep_links=False
	)

	if job["tabs"] == "":
		data = read_xlsx_sheet(wb[wb.get_sheet_names()[0]], data, job)
		ntabs = 1
	else:
		data["headers"].append(job["tabs"])
		sheets = wb.get_sheet_names()
		if job["skip_tabs"] == "":
			job["skip_tabs"] = 0
		for i in range(int(job["skip_tabs"]), len(sheets)):
			print_if_verbose("Reading sheet: '" + sheets[i] + "'")
			data = read_xlsx_sheet(wb[sheets[i]], data, job, sheets[i])
			ntabs += 1
	wb.close()
	print_with_timestamp(str(ntabs) + " tab[s] read")
	return data



def read_xlsx_sheet(sheet, data, job, sheetname=""):
	header = int(job["header"])
	if job["subheader"] == "":
		subheader = None
		firstrow = int(job["header"]) + 1
	else:
		subheader = int(job["subheader"])
		firstrow = int(job["subheader"]) + 1

	if job["column_wrap"] == "":
		ncols = sheet.max_column
		nframes = 1
	else:
		ncols = int(job["column_wrap"])
		if sheet.max_column % ncols > 0:
			nframes = sheet.max_column // ncols + 1
		else:
			nframes = sheet.max_column // ncols
	prev_head = ""
	col_names = []

	for col in range(1, ncols + 1): # openpyxl is 1-indexed
		if subheader is None:
			val = sheet.cell(row=header, column=col).value
		else:
			if sheet.cell(row=header, column=col).value != None:
				prev_head = sheet.cell(row=header, column=col).value
			if sheet.cell(row=subheader, column=col).value == None:
				val = prev_head
			else:
				val = prev_head + ": " + str(sheet.cell(row=subheader, column=col).value)
		col_names.append(val)

	for val in col_names:
		if val is not None and val not in data["headers"]:
			data["headers"].append(val)

	for frame in range(0, nframes):
		for rownum in range(firstrow, sheet.max_row + 1):
			content = {}
			if job["tabs"] != "":
				content[job["tabs"]] = sheetname
			for col in range(0, len(col_names)):
				if col_names[col] is not None:
					content[col_names[col]] = clean_value(sheet.cell(row=rownum, column=col+frame*ncols+1).value, True)
			data["rows"].append(content)

	return data





def write_csv(data, filename):
	with open(filename, 'w') as outfile:
		writer = csv.DictWriter(outfile, fieldnames=data["headers"])
		writer.writeheader()
		for row in data["rows"]:
			writer.writerow(row)



# "turkish" in this a special case for routine ways Turkish characters:
# İ, Ṣ, ğ, ı & ṣ
# get mangled
def clean_value(val, turkish=False):
	if val is None:
		return ""
	elif turkish:
		return str(val).replace('Ý','İ').replace('Þ', 'Ş').replace('ð','ğ').replace('ý', 'ı').replace('þ', 'ş')
	else:
		return val





def get_args():
	parser = argparse.ArgumentParser(description="Make flat CSVs out of tabbed Excel files")

# positional argument
	parser.add_argument("joblist", help="required argument: list of files to process with some metadata described in comments at the top of the script.")

	args = parser.parse_args()
	return args



def print_if_verbose(msg):
	if verbose:
		if msg == "":
			print(" ")
		else:
			print_with_timestamp(msg)



def print_with_timestamp(msg):
	print(time.ctime() + ": " + str(msg))
	sys.stdout.flush()
# explicitly flushing stdout makes sure that a .out file stays up to date
# otherwise it can be hard to keep track of whether a background job is hanging




def elapsed_time(starttime):
	seconds = time.time() - starttime
	if seconds < 1:
		return "less than one second"
	hours = int(seconds / 60 / 60)
	minutes = int(seconds / 60 - hours * 60)
	seconds = int(seconds - minutes * 60 - hours * 60 * 60)
	if minutes < 1 and hours < 1:
		return str(seconds) + " seconds"
	elif hours < 1:
		return str(minutes) + " minute[s] and " + str(seconds) + " second[s]"
	else:
		return (
				str(hours) + " hour[s], " +
				str(minutes) + " minute[s] and " +
				str(seconds) + " second[s]"
		)


if __name__ == "__main__":
	try:
		main()
	except:
		import sys
		print(sys.exc_info()[0])
		import traceback
		print(traceback.format_exc())
	if verbose:
		print("Press Enter to continue ...")
		input()
